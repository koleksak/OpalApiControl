"""Excel file parser for ePhasorsim Models"""

import logging
logging.basicConfig(level=logging.INFO)
import os
import re
from openpyxl import load_workbook
from settings import Settings
from numpy import long


def excel_parse(file):
    """Creates Sys Parameter Dictionary based on an excel file containing model initialization data for 
    ePHASORsim."""
    # TODO: Check data device param ordering with PSS/E initialization data parser.
    ExcelParams = {}
    logging.info("<Loading Excel file>")
    wb = load_workbook(file, read_only=True)  #Loads excel file as a workbook with openpyxl methods for parsing
    logging.info("<Parsing Excel file>")
    for sheet in wb.get_sheet_names():#TODO: Add sub params for these non-data sheets..Transformer Exception
        if sheet == 'General':
            ExcelParams[sheet] = {}
            ExcelParams[sheet]['Settings'] = {}
            current_sheet = wb[sheet]
            for rows in range(1,current_sheet.max_row):
                settingName = current_sheet._get_cell(rows,2)
                settingData = current_sheet._get_cell(rows, 2)
                if settingData.value == None:
                    break
                else:
                    value = to_basic_types(settingData.value)
                    ExcelParams[sheet]['Settings'][settingName.value] = value

        elif sheet == 'Pins':
            continue

        else:
            print sheet
            ExcelParams[sheet] = {}
            ExcelParams[sheet]['ParamsList'] = []
            current_sheet = wb[sheet]
            #Adds parameter to Excel Dictionary
            for index in range(1,current_sheet.max_column):
                param = current_sheet._get_cell(1, index)
                if param.value == None:
                    ExcelParams[sheet]['cols'] = index - 1
                    break
                else:
                    ExcelParams[sheet]['ParamsList'].append(param.value)
                    ExcelParams[sheet][param.value] = []

            #Adds values to parameter dictionary after extracting number/names of params
            #convert_to_int_list = ['Bus', 'From bus', 'To bus']
            for col in range(1, ExcelParams[sheet]['cols']+1):
                for row in range(2, current_sheet.max_row):
                    param_val = current_sheet._get_cell(row, col)
                    if param_val.value == None:
                        break
                    else:
                        param_name = ExcelParams[sheet]['ParamsList'][col-1]
                        value = to_basic_types(param_val.value)
                        ExcelParams[sheet][param_name].append(value)

    return ExcelParams  #TODO: Set to SysParam Class equivalent


def to_basic_types(data):
    """Convert data to the basic Python types (int, float)"""

    retval = None
    if type(data) == unicode:
        retval = data
    if type(data) == long:
        data = float(data)
    if type(data) == float and data.is_integer():
        retval = int(data)
    else:
        retval = data

    return retval



if __name__ == "__main__":
    file = 'C:/Users/opalrt/repos/OpalApiControl/OpalApiControl/tests/phasor01_IEEE39.xlsx'
    Params = excel_parse(file)
    print Params
